import matlab.engine
import pandas as pd
import openpyxl
from openpyxl import Workbook

file_name = "E:/RASLAB/PROJECT_RUN_THROUGH/Project_parameters_file.xlsx"

def save_dataframe_to_excel(file_name, sheet_name, dataframe):
    """
    Saves a DataFrame to an Excel file, replacing the specified sheet if it already exists.

    :param file_name: The name of the Excel file.
    :param sheet_name: The name of the sheet to save the DataFrame to.
    :param dataframe: The pandas DataFrame to save.
    """
    try:
        # Attempt to load the existing workbook, if it exists
        try:
            workbook = openpyxl.load_workbook(file_name)
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook
            workbook = Workbook()
        
        # If sheet already exists, delete it
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        
        # Create a new sheet
        sheet = workbook.create_sheet(sheet_name)

        # Write DataFrame headers to Excel
        for col_idx, col_name in enumerate(dataframe.columns, start=1):
            sheet.cell(row=1, column=col_idx, value=col_name)

        # Write DataFrame rows to Excel
        for row_idx, row in enumerate(dataframe.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Save workbook
        workbook.save(file_name)
        print(f"The DataFrame has been saved to the '{sheet_name}' sheet in {file_name}.")
    
    except Exception as e:
        print(f"An error occurred while saving DataFrame to Excel: {e}")
        
def read_3d_coordinates(sheet_name, file_name):
    """
    Reads 3D coordinates (x, y, z) from the specified sheet in the Excel file.

    :param sheet_name: The name of the sheet to read data from.
    :param file_name: The name of the Excel file to read the data from.
    :return: A list of lists, each containing [x, y, z] from each row in the Excel sheet.
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_name)

        # Check if the specified sheet exists
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"The sheet '{sheet_name}' does not exist in {file_name}.")

        # Access the desired sheet
        sheet = workbook[sheet_name]

        # Read data into a list of lists
        coordinates = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=4, values_only=True):
            coordinates.append([row[0], row[1], row[2]])

        print(f"Successfully read 3D coordinates from '{sheet_name}' in {file_name}.")
        return coordinates

    except Exception as e:
        print(f"An error occurred while reading 3D coordinates: {e}")
        return []


# Start the MATLAB engine
eng = matlab.engine.start_matlab()

# Load target poses from CSV file and Convert the DataFrame to a Python list
target_pose = read_3d_coordinates("Physical_coordinates", file_name)
#print (target_pose)

# Convert Python list to MATLAB matrix
matlab_pose = matlab.double(target_pose)

# Call the MATLAB function to compute joint angles
ik_angles = eng.my_kinematics('ik', matlab_pose)

# Convert MATLAB output to Python format
ik_angles = [list(row) for row in ik_angles]

# Save the output angles to a CSV file
angles_df = pd.DataFrame(ik_angles, columns=[f'Joint{i+1}' for i in range(len(ik_angles[0]))])
print(angles_df)
save_dataframe_to_excel(file_name, "Joint_Angles", angles_df)

# Stop the MATLAB engine
eng.quit()
