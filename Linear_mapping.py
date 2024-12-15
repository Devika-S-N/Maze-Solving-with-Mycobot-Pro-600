import csv
import os
import openpyxl
from openpyxl import Workbook

file_name = "E:/RASLAB/PROJECT_RUN_THROUGH/Project_parameters_file.xlsx"

def read_calibration_parameters():
    """
    Reads the calibration parameters (mx, cx, my, cy) from the 'calibration_parameters' 
    sheet in the 'Project_parameters_file.xlsx' file.

    Returns:
        dict: A dictionary containing the calibration parameters with keys 'mx', 'cx', 'my', 'cy'.
    """
    global file_name
    sheet_name = "calibration_parameters"

    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        print(f"Error: {file_name} does not exist.")
        return {}

    # Check if the 'calibration_parameters' sheet exists
    if sheet_name not in workbook.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found in {file_name}.")
        return {}

    # Access the sheet
    sheet = workbook[sheet_name]

    # Initialize dictionary to store parameters
    calibration_parameters = {}

    # Read values from the sheet
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
        key, value = row
        if key in ['mx', 'cx', 'my', 'cy']:  # Only process expected keys
            calibration_parameters[key] = value

    return calibration_parameters

def read_pixel_coordinates():
    """
    Reads the 'Pixel_coordinates' sheet from the 'Project_parameters_file.xlsx' Excel file
    and returns the list of points in the format [(x1, y1), (x2, y2), ...].

    Returns:
        list of tuples: List of (x, y) coordinates.
    """
    global file_name
    sheet_name = "Pixel_coordinates"
    
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        print(f"Error: {file_name} does not exist.")
        return []
    
    # Check if the 'Pixel_coordinates' sheet exists
    if sheet_name not in workbook.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found in {file_name}.")
        return []
    
    # Access the sheet
    sheet = workbook[sheet_name]
    
    # Read the points from the sheet
    points = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3, values_only=True):
        x, y = row
        if x is not None and y is not None:  # Ensure non-empty cells
            points.append((x, y))
    
    return points

def getRobotCoordinates(cameraX, cameraY, mx, cx, my, cy):
    """
    Converts camera coordinates (cameraX, cameraY) to robot physical coordinates.
    :param cameraX: Camera X-coordinate
    :param cameraY: Camera Y-coordinate
    :param mx: X-axis scaling factor
    :param cx: X-axis offset
    :param my: Y-axis scaling factor
    :param cy: Y-axis offset
    :return: Physical coordinates (x, y) in robot space
    """
    return (cameraX * mx + cx, cameraY * my + cy)


def add_pixel_coordinates(file_name, sheet_name, aligned_points):
    """
    Adds or refreshes the 'Pixel_coordinates' sheet with given aligned points.

    :param file_name: The name of the Excel file.
    :param sheet_name: The name of the sheet to create or refresh.
    :param aligned_points: List of tuples containing x and y coordinates [(x1, y1), (x2, y2), ...].
    """
    try:
        # Load existing workbook or create a new one
        try:
            workbook = openpyxl.load_workbook(file_name)
        except FileNotFoundError:
            workbook = Workbook()

        # Check if the sheet already exists
        if sheet_name in workbook.sheetnames:
            # Remove the existing sheet
            del workbook[sheet_name]

        # Add a new sheet
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(["Point", "X", "Y"])
        # Write aligned points to the sheet
        for index, (x, y) in enumerate(aligned_points, start=1):
            sheet.append([f"point {index}", x, y])

        # Save the workbook
        workbook.save(file_name)
        print(f"The '{sheet_name}' sheet has been updated in {file_name}.")
    except Exception as e:
        print(f"An error occurred while adding pixel coordinates: {e}")
        

def add_physical_coordinates(file_name, sheet_name, aligned_points):
    """
    Adds or refreshes the specified sheet with given 3D points (x, y, z).

    :param sheet_name: The name of the Excel sheet.
    :param aligned_points: List of tuples containing (x, y, z) coordinates [(x1, y1, z1), (x2, y2, z2), ...].
    :param file_name: The name of the Excel file to save the data into.
    """
    try:
        # Load existing workbook or create a new one
        try:
            workbook = openpyxl.load_workbook(file_name)
        except FileNotFoundError:
            workbook = Workbook()

        # Check if the sheet already exists
        if sheet_name in workbook.sheetnames:
            # Remove the existing sheet
            del workbook[sheet_name]

        # Add a new sheet
        sheet = workbook.create_sheet(sheet_name)

        # Add the headers
        sheet.append(["Point", "PhysicalX", "PhysicalY", "PhysicalZ"])

        # Write aligned points to the sheet
        for index, (x, y, z) in enumerate(aligned_points, start=1):
            sheet.append([f"{index}", x, y, z])

        # Save the workbook
        workbook.save(file_name)
        print(f"The '{sheet_name}' sheet has been updated in {file_name}.")
    
    except Exception as e:
        print(f"An error occurred while adding 3D coordinates: {e}")



def main():
    # Step 1: Get the calibration parameters
    parameters = read_calibration_parameters()
    mx = parameters['mx']
    my = parameters['my']
    cx = parameters['cx']
    cy = parameters['cy']
    physicalZ = 0.07

    # Step 2: Get the x and y coordinates for each point from the file
    points = read_pixel_coordinates()

    # Step 3: Calculate physical coordinates for each point
    physical_coordinates = []
    print("\nPhysical coordinates of the points:")
    for i, (cameraX, cameraY) in enumerate(points):
        physicalX, physicalY = getRobotCoordinates(cameraX, cameraY, mx, cx, my, cy)
        physicalX = physicalX * 0.001  # Convert to desired units (mm to meters, for example)
        physicalY = physicalY * 0.001
        print(f"Point {i+1}: Camera Coordinates ({cameraX}, {cameraY}) -> Physical Coordinates ({physicalX}, {physicalY})")
        physical_coordinates.append((physicalX, physicalY, physicalZ))

    # Step 4: Save the physical coordinates to a CSV file
    add_physical_coordinates(file_name,"Physical_coordinates",physical_coordinates)
    

if __name__ == "__main__":
    main()
